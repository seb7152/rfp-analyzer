#!/usr/bin/env python3
"""
Template: Extract Supplier Responses from Excel File

This is a template for extracting supplier responses from Excel files where
responses are organized as columns (one column per supplier).

CUSTOMIZE THIS TEMPLATE:
1. Update the sheet name if different
2. Specify which suppliers to extract
3. Adjust column indices to match your file
4. Set up score mapping if applicable
5. Test with sample data before running on full file

File structure: Sheet with columns:
                Requirement Code, Requirement Title, ..., Supplier A, Supplier B, Supplier C
Created: 2024-01-10

Usage:
    python extract_responses_supplier_columns.py input.xlsx output.json
    python extract_responses_supplier_columns.py input.xlsx output.json "Supplier A,Supplier B"

Optionally specify which suppliers to extract (comma-separated).
"""

import json
import sys
import openpyxl
from pathlib import Path


def extract_supplier_responses(file_path, supplier_names=None):
    """
    Extract supplier responses from Excel file where suppliers are in columns.

    Args:
        file_path: Path to Excel file
        supplier_names: List of supplier column names to extract. If None, auto-detect.

    Returns:
        List of response dictionaries matching the JSON schema
    """
    print(f"Loading workbook: {file_path}")

    try:
        wb = openpyxl.load_workbook(file_path)
    except FileNotFoundError:
        print(f"ERROR: File not found: {file_path}")
        sys.exit(1)
    except Exception as e:
        print(f"ERROR: Failed to load workbook: {e}")
        sys.exit(1)

    # ============================================================================
    # CUSTOMIZE: Update sheet name to match your file
    # ============================================================================
    sheet_name = "Responses"  # Change this if your sheet has a different name

    if sheet_name not in wb.sheetnames:
        print(f"ERROR: Sheet '{sheet_name}' not found.")
        print(f"Available sheets: {wb.sheetnames}")
        sys.exit(1)

    ws = wb[sheet_name]
    print(f"Using sheet: {sheet_name}")

    # ============================================================================
    # CUSTOMIZE: Update column indices
    # ============================================================================
    REQUIREMENT_CODE_COL = 0      # Column A - requirement code identifier
    REQUIREMENT_TITLE_COL = 1     # Column B - requirement title (optional)
    SUPPLIER_START_COL = 2        # Column C onwards - supplier responses start here

    HEADER_ROW = 1
    DATA_START_ROW = 2

    # ============================================================================
    # Extract header to identify supplier columns
    # ============================================================================
    print("Reading header row to identify suppliers...")
    header_row = ws[HEADER_ROW]
    all_headers = [cell.value for cell in header_row]

    supplier_columns = {}  # {supplier_name: column_index}

    # Auto-detect or filter suppliers
    for col_idx in range(SUPPLIER_START_COL, len(all_headers)):
        header = all_headers[col_idx]
        if header:
            supplier_name = str(header).strip()

            # Filter by requested suppliers if specified
            if supplier_names and supplier_name not in supplier_names:
                continue

            supplier_columns[supplier_name] = col_idx
            print(f"  Found supplier: {supplier_name} (column {col_idx + 1})")

    if not supplier_columns:
        print("ERROR: No suppliers found. Check SUPPLIER_START_COL and supplier names.")
        sys.exit(1)

    print(f"\nExtracting responses for {len(supplier_columns)} supplier(s)")

    # ============================================================================
    # Extract responses
    # ============================================================================
    data = []
    skipped_count = 0

    print(f"Extracting from row {DATA_START_ROW} onwards...")

    for row_num, row in enumerate(ws.iter_rows(min_row=DATA_START_ROW, values_only=False),
                                   start=DATA_START_ROW):

        # Get requirement code
        requirement_code = row[REQUIREMENT_CODE_COL].value if REQUIREMENT_CODE_COL < len(row) else None

        # Skip empty rows
        if not requirement_code or (isinstance(requirement_code, str) and not requirement_code.strip()):
            continue

        requirement_code = str(requirement_code).strip()

        # ====================================================================
        # Extract response for each supplier
        # ====================================================================
        for supplier_name, col_idx in supplier_columns.items():

            response_text = row[col_idx].value if col_idx < len(row) else None

            # Skip empty responses
            if response_text is None:
                continue

            response_text = str(response_text).strip() if response_text else None

            if not response_text:
                continue

            # ================================================================
            # CUSTOMIZE: Add score extraction if available
            # ================================================================
            # Example: If scores are in columns, extract them here
            # score_col = col_idx + 1  # Score in next column
            # score = row[score_col].value if score_col < len(row) else None
            # Try to convert to float (0-5 range)

            ai_score = None
            # Uncomment and customize if scores are available:
            # if score is not None:
            #     try:
            #         ai_score = float(score)
            #         if not (0 <= ai_score <= 5):
            #             ai_score = None
            #     except (ValueError, TypeError):
            #         pass

            # ================================================================
            # Build response object
            # ================================================================
            response = {
                "requirement_id_external": requirement_code,
                "response_text": response_text,
            }

            # Add optional fields if present
            if ai_score is not None:
                response["ai_score"] = ai_score

            # Optional: Add supplier name as comment (helps identify source)
            # response["ai_comment"] = f"Response from {supplier_name}"

            data.append(response)

        # Log progress
        if row_num % 50 == 0:
            print(f"  Processed row {row_num}...")

    print(f"Extracted {len(data)} responses (skipped {skipped_count})")

    return data


def parse_supplier_list(supplier_string):
    """Parse comma-separated supplier names"""
    if not supplier_string:
        return None
    return [s.strip() for s in supplier_string.split(",") if s.strip()]


def main():
    """Main entry point"""
    if len(sys.argv) < 3:
        print(f"Usage: python {sys.argv[0]} <input_file.xlsx> <output.json> [suppliers]")
        print()
        print("Arguments:")
        print("  input_file.xlsx    : Path to Excel file with supplier responses")
        print("  output.json        : Output JSON file path")
        print("  suppliers          : Optional comma-separated list of suppliers to extract")
        print("                      If not specified, all suppliers will be extracted")
        print()
        print("Examples:")
        print(f"  python {sys.argv[0]} responses.xlsx output.json")
        print(f"  python {sys.argv[0]} responses.xlsx output.json 'Supplier A,Supplier B'")
        sys.exit(1)

    input_file = sys.argv[1]
    output_file = sys.argv[2]
    supplier_filter = parse_supplier_list(sys.argv[3]) if len(sys.argv) > 3 else None

    print("=" * 70)
    print("Supplier Response Extractor - Column-Based Template")
    print("=" * 70)
    print()

    if supplier_filter:
        print(f"Filtering for suppliers: {', '.join(supplier_filter)}")
        print()

    # Extract
    responses = extract_supplier_responses(input_file, supplier_filter)

    if not responses:
        print("ERROR: No responses extracted. Check file structure and supplier columns.")
        sys.exit(1)

    # Save to JSON
    print(f"\nSaving to JSON: {output_file}")
    try:
        with open(output_file, 'w', encoding='utf-8') as f:
            json.dump(responses, f, indent=2, ensure_ascii=False)
        print("✓ Successfully saved")
    except Exception as e:
        print(f"ERROR: Failed to save JSON: {e}")
        sys.exit(1)

    # Summary
    print()
    print("Summary:")
    print(f"  - Total responses: {len(responses)}")

    if responses:
        # Count unique requirements
        unique_reqs = set(r['requirement_id_external'] for r in responses)
        print(f"  - Unique requirements: {len(unique_reqs)}")

        # Count responses per requirement
        from collections import Counter
        req_counts = Counter(r['requirement_id_external'] for r in responses)
        max_responses = max(req_counts.values())
        print(f"  - Max responses per requirement: {max_responses}")

        # Show first example
        print(f"  - First response: {responses[0]['requirement_id_external']} -> \"{responses[0]['response_text'][:50]}...\"")

    print()
    print("Next steps:")
    print(f"  1. Review the output: cat {output_file} | head -n 50")
    print(f"  2. Validate the JSON: python validate_json.py {output_file} responses")
    print(f"  3. Import to RFP Analyzer system")


if __name__ == "__main__":
    main()
