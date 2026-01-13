#!/usr/bin/env python3
"""
Template: Extract Requirements from Simple Excel File

This is a template for extracting requirements from Excel files with a simple
table structure (columns with headers, data starting at row 2).

CUSTOMIZE THIS TEMPLATE:
1. Update the sheet name if different from "Requirements"
2. Adjust column indices to match your file
3. Add/remove field mappings as needed
4. Test with sample data before running on full file

Original file: [UPDATE THIS]
File structure: Single sheet with columns:
                Code, Title, Description, Weight, Category
Created: 2024-01-10

Usage:
    python extract_requirements_simple_excel.py input.xlsx requirements.json

Adjust column mappings and sheet name before running.
"""

import json
import sys
import openpyxl
from pathlib import Path


def extract_requirements(file_path):
    """
    Extract requirements from Excel file with simple column structure.

    Returns:
        List of requirement dictionaries matching the JSON schema
    """
    print(f"Loading workbook: {file_path}")

    try:
        wb = openpyxl.load_workbook(file_path)
    except FileNotFoundError:
        print(f"ERROR: File not found: {file_path}")
        sys.exit(1)
    except Exception as e:
        print(f"ERROR: Failed to load workbook: {e}")
        sys.exit(1)

    # ============================================================================
    # CUSTOMIZE: Update sheet name to match your file
    # ============================================================================
    sheet_name = "Requirements"  # Change this if your sheet has a different name

    if sheet_name not in wb.sheetnames:
        print(f"ERROR: Sheet '{sheet_name}' not found.")
        print(f"Available sheets: {wb.sheetnames}")
        sys.exit(1)

    ws = wb[sheet_name]
    print(f"Using sheet: {sheet_name}")

    # ============================================================================
    # CUSTOMIZE: Update column indices to match your file structure
    # Column indices are 0-based: A=0, B=1, C=2, D=3, E=4, etc.
    # ============================================================================
    CODE_COL = 0           # Column A - requirement code
    TITLE_COL = 1          # Column B - requirement title
    DESCRIPTION_COL = 2    # Column C - detailed description
    WEIGHT_COL = 3         # Column D - weight/importance (0-1 or percentage)
    CATEGORY_COL = 4       # Column E - category name
    TAGS_COL = None        # Column ? - optional tags (set to None if not present)
    MANDATORY_COL = None   # Column ? - optional is_mandatory flag
    PAGE_NUM_COL = None    # Column ? - optional page number

    # Header row (usually row 1, but adjust if different)
    HEADER_ROW = 1
    DATA_START_ROW = 2     # First row with data

    data = []
    skipped_count = 0

    print(f"Extracting from row {DATA_START_ROW} onwards...")

    # ============================================================================
    # Main extraction loop
    # ============================================================================
    for row_num, row in enumerate(ws.iter_rows(min_row=DATA_START_ROW, values_only=False),
                                   start=DATA_START_ROW):

        # Get values from columns
        code = row[CODE_COL].value if CODE_COL < len(row) else None
        title = row[TITLE_COL].value if TITLE_COL < len(row) else None
        description = row[DESCRIPTION_COL].value if DESCRIPTION_COL < len(row) else None
        weight_raw = row[WEIGHT_COL].value if WEIGHT_COL < len(row) else None
        category = row[CATEGORY_COL].value if CATEGORY_COL < len(row) else None

        # Optional fields
        tags_raw = row[TAGS_COL].value if TAGS_COL is not None and TAGS_COL < len(row) else None
        is_mandatory = row[MANDATORY_COL].value if MANDATORY_COL is not None and MANDATORY_COL < len(row) else None
        page_num = row[PAGE_NUM_COL].value if PAGE_NUM_COL is not None and PAGE_NUM_COL < len(row) else None

        # Skip empty rows
        if not code or (isinstance(code, str) and not code.strip()):
            continue

        # ====================================================================
        # CUSTOMIZE: Add your own data cleaning and validation logic here
        # ====================================================================

        # Strip whitespace from strings
        code = str(code).strip() if code else None
        title = str(title).strip() if title else None
        description = str(description).strip() if description else None
        category = str(category).strip() if category else None

        # Validate required fields
        if not code or not title or not description:
            print(f"WARNING: Row {row_num} - missing required fields (code, title, description). Skipping.")
            skipped_count += 1
            continue

        # Convert weight to float
        weight = None
        if weight_raw is not None:
            try:
                weight_val = float(weight_raw)
                # If weight is in percentage (>1), convert to 0-1 range
                if weight_val > 1:
                    weight = weight_val / 100
                else:
                    weight = weight_val

                # Validate weight range
                if not (0 <= weight <= 1):
                    print(f"WARNING: Row {row_num} - weight {weight_val} is outside 0-1 range. Using default 0.5")
                    weight = 0.5
            except (ValueError, TypeError):
                print(f"WARNING: Row {row_num} - could not convert weight '{weight_raw}' to number. Using default 0.5")
                weight = 0.5
        else:
            weight = 0.5  # Default weight if not provided

        # Handle category - use provided value or default
        if not category:
            category = "General"  # Default category if not specified

        # Parse tags (comma-separated)
        tags = []
        if tags_raw:
            tags = [tag.strip() for tag in str(tags_raw).split(",") if tag.strip()]

        # Convert boolean fields
        if isinstance(is_mandatory, str):
            is_mandatory = is_mandatory.lower() in ['true', 'yes', '1', 'x']
        elif is_mandatory is None:
            is_mandatory = False

        # Convert page number to integer
        page_number = None
        if page_num is not None:
            try:
                page_number = int(page_num)
                if page_number < 1:
                    page_number = None
            except (ValueError, TypeError):
                pass

        # ====================================================================
        # Build requirement object matching JSON schema
        # ====================================================================
        requirement = {
            "code": code,
            "title": title,
            "description": description,
            "weight": weight,
            "category_name": category,
        }

        # Add optional fields only if present
        if tags:
            requirement["tags"] = tags
        if is_mandatory:
            requirement["is_mandatory"] = is_mandatory
        if page_number:
            requirement["page_number"] = page_number

        data.append(requirement)

        # Log progress
        if len(data) % 50 == 0:
            print(f"  Processed {len(data)} requirements...")

    print(f"Extracted {len(data)} requirements (skipped {skipped_count})")

    return data


def main():
    """Main entry point"""
    if len(sys.argv) != 3:
        print(f"Usage: python {sys.argv[0]} <input_file.xlsx> <output.json>")
        print()
        print("Examples:")
        print(f"  python {sys.argv[0]} requirements.xlsx requirements.json")
        print(f"  python {sys.argv[0]} rfp_data.xlsx extracted.json")
        sys.exit(1)

    input_file = sys.argv[1]
    output_file = sys.argv[2]

    print("=" * 70)
    print("RFP Requirement Extractor - Simple Excel Template")
    print("=" * 70)
    print()

    # Extract
    requirements = extract_requirements(input_file)

    if not requirements:
        print("ERROR: No requirements extracted. Check file structure and column mappings.")
        sys.exit(1)

    # Save to JSON
    print(f"\nSaving to JSON: {output_file}")
    try:
        with open(output_file, 'w', encoding='utf-8') as f:
            json.dump(requirements, f, indent=2, ensure_ascii=False)
        print("✓ Successfully saved")
    except Exception as e:
        print(f"ERROR: Failed to save JSON: {e}")
        sys.exit(1)

    # Summary
    print()
    print("Summary:")
    print(f"  - Total requirements: {len(requirements)}")
    if requirements:
        print(f"  - First requirement: {requirements[0]['code']} - {requirements[0]['title']}")
        categories = set(req.get('category_name') for req in requirements)
        print(f"  - Categories found: {len(categories)}")

    print()
    print("Next steps:")
    print(f"  1. Review the output: cat {output_file} | head -n 20")
    print(f"  2. Validate the JSON: python validate_json.py {output_file} requirements")
    print(f"  3. Import to RFP Analyzer system")


if __name__ == "__main__":
    main()
